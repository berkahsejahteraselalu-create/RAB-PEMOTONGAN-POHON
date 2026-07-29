import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime

def create_styled_rab_excel(output_path):
    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # STYLES & PALETTE DEFINITIONS (Emerald Executive Theme)
    # ---------------------------------------------------------
    FONT_NAME = "Segoe UI"
    
    # Color Codes
    HEX_DARK_EMERALD  = "1E4D2B"
    HEX_MEDIUM_FOREST = "2D6A4F"
    HEX_LIGHT_SAGE    = "D8F3DC"
    HEX_MINT_TINT     = "E9F5EC"
    HEX_ZEBRA_LIGHT   = "F8FBF9"
    HEX_CARD_BG       = "F4F9F5"
    HEX_BORDER_GRAY   = "D0D7D1"
    HEX_TEXT_DARK     = "081C15"
    
    # Fonts
    font_title_main  = Font(name=FONT_NAME, size=15, bold=True, color="FFFFFF")
    font_subtitle    = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    font_reg_italic  = Font(name=FONT_NAME, size=9.5, italic=True, color=HEX_TEXT_DARK)
    font_col_hdr     = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    font_sec_hdr     = Font(name=FONT_NAME, size=11, bold=True, color=HEX_TEXT_DARK)
    font_item_hdr    = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    font_sub_grp     = Font(name=FONT_NAME, size=9.5, bold=True, color="1B4332")
    font_data        = Font(name=FONT_NAME, size=9.5, color="000000")
    font_data_bold   = Font(name=FONT_NAME, size=9.5, bold=True, color="000000")
    font_subtotal    = Font(name=FONT_NAME, size=9.5, bold=True, color="1B4332")
    font_direct_cost = Font(name=FONT_NAME, size=10, bold=True, color=HEX_TEXT_DARK)
    font_overhead    = Font(name=FONT_NAME, size=9.5, italic=True, color="2D6A4F")
    font_grand_unit  = Font(name=FONT_NAME, size=10.5, bold=True, color="FFFFFF")
    font_grand_total = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    font_card_lbl    = Font(name=FONT_NAME, size=9.5, bold=True, color="1B4332")
    font_card_val    = Font(name=FONT_NAME, size=9.5, color="000000")

    # Fills
    fill_dark_emerald  = PatternFill(start_color=HEX_DARK_EMERALD, end_color=HEX_DARK_EMERALD, fill_type="solid")
    fill_medium_forest = PatternFill(start_color=HEX_MEDIUM_FOREST, end_color=HEX_MEDIUM_FOREST, fill_type="solid")
    fill_light_sage    = PatternFill(start_color=HEX_LIGHT_SAGE, end_color=HEX_LIGHT_SAGE, fill_type="solid")
    fill_mint_tint     = PatternFill(start_color=HEX_MINT_TINT, end_color=HEX_MINT_TINT, fill_type="solid")
    fill_zebra_light   = PatternFill(start_color=HEX_ZEBRA_LIGHT, end_color=HEX_ZEBRA_LIGHT, fill_type="solid")
    fill_card_bg       = PatternFill(start_color=HEX_CARD_BG, end_color=HEX_CARD_BG, fill_type="solid")

    # Borders
    bdr_thin_gray = Side(style="thin", color=HEX_BORDER_GRAY)
    bdr_med_green = Side(style="medium", color=HEX_MEDIUM_FOREST)
    bdr_dbl_green = Side(style="double", color=HEX_DARK_EMERALD)
    
    border_data      = Border(left=bdr_thin_gray, right=bdr_thin_gray, top=bdr_thin_gray, bottom=bdr_thin_gray)
    border_hdr       = Border(left=bdr_med_green, right=bdr_med_green, top=bdr_med_green, bottom=bdr_med_green)
    border_subtotal  = Border(left=bdr_thin_gray, right=bdr_thin_gray, top=bdr_thin_gray, bottom=bdr_thin_gray)
    border_total     = Border(left=bdr_thin_gray, right=bdr_thin_gray, top=bdr_thin_gray, bottom=bdr_dbl_green)
    border_grand     = Border(left=bdr_med_green, right=bdr_med_green, top=bdr_med_green, bottom=bdr_dbl_green)
    border_card      = Border(left=bdr_med_green, right=bdr_thin_gray, top=bdr_thin_gray, bottom=bdr_thin_gray)

    # Alignments
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left_wrap   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_left        = Alignment(horizontal="left", vertical="center")
    align_right       = Alignment(horizontal="right", vertical="center")
    align_center      = Alignment(horizontal="center", vertical="center")

    # Number Formats
    FMT_RP   = '_("Rp "* #,##0_);_("Rp "* (#,##0);_("Rp "* "-"_);_(@_)'
    FMT_COEF = '0.000'
    FMT_INT  = '#,##0'

    # Helper function to style merged range
    def style_range(ws, cell_range, font=None, fill=None, alignment=None, border=None):
        for row in ws[cell_range]:
            for cell in row:
                if font: cell.font = font
                if fill: cell.fill = fill
                if alignment: cell.alignment = alignment
                if border: cell.border = border

    # =========================================================================
    # SHEET 1: COVER
    # =========================================================================
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_cover.merge_cells("A1:F1")
    ws_cover["A1"] = "RENCANA ANGGARAN BIAYA (RAB)"
    style_range(ws_cover, "A1:F1", font=font_title_main, fill=fill_dark_emerald, alignment=align_center_wrap, border=border_hdr)
    ws_cover.row_dimensions[1].height = 36

    ws_cover.merge_cells("A2:F2")
    ws_cover["A2"] = "PEKERJAAN PEMOTONGAN / PERAMPINGAN POHON"
    style_range(ws_cover, "A2:F2", font=font_subtitle, fill=fill_medium_forest, alignment=align_center_wrap)
    ws_cover.row_dimensions[2].height = 24

    ws_cover.merge_cells("A3:F3")
    ws_cover["A3"] = "Berdasarkan SE Ditjen Bina Konstruksi No. 47/SE/Dk/2026 (AHSP PUPR 2026)"
    style_range(ws_cover, "A3:F3", font=font_reg_italic, fill=fill_light_sage, alignment=align_center_wrap)
    ws_cover.row_dimensions[3].height = 20

    ws_cover.row_dimensions[4].height = 12

    # Information Card Header
    ws_cover.merge_cells("A5:B5")
    ws_cover["A5"] = "📌 INFORMASI PROYEK & REGULASI"
    style_range(ws_cover, "A5:B5", font=font_sub_grp, fill=fill_mint_tint, alignment=align_left, border=border_data)

    ws_cover.merge_cells("D5:F5")
    ws_cover["D5"] = "📊 RINGKASAN ESTIMASI BIAYA (SUMMARY)"
    style_range(ws_cover, "D5:F5", font=font_sub_grp, fill=fill_mint_tint, alignment=align_left, border=border_data)

    infos = [
        ("Nama Pekerjaan", "Pemeliharaan Pohon Jalan & Ruang Terbuka Hijau (RTH)"),
        ("Nomor Dokumen", "RAB-PPH-2026-001"),
        ("Tahun Anggaran", "2026"),
        ("Tanggal RAB", datetime.datetime.now().strftime("%d %B %Y")),
        ("Satuan Biaya", "Rupiah (Rp)"),
        ("Dasar Acuan", "SE Ditjen Bina Konstruksi No. 47/SE/Dk/2026"),
        ("Biaya SMKK (K3)", "2,5% dari Biaya Fisik (Permen PUPR 10/2021)"),
        ("Pajak (PPN)", "11% (UU HPP No. 7/2021)"),
        ("Overhead & Profit", "15% dari Biaya Langsung"),
        ("Standar Ukur DBH", "DBH = 1,30 meter dari permukaan tanah"),
    ]

    summary_items = [
        ("Sub-Total Perampingan (Kelompok A)", "='RAB Rekapitulasi'!G11"),
        ("Sub-Total Pemotongan (Kelompok B)", "='RAB Rekapitulasi'!G19"),
        ("Total Biaya Langsung Fisik", "='RAB Rekapitulasi'!G20"),
        ("Biaya Penerapan SMKK (2.5%)", "='RAB Rekapitulasi'!G21"),
        ("Jumlah Fisik + SMKK", "='RAB Rekapitulasi'!G22"),
        ("PPN (11%)", "='RAB Rekapitulasi'!G23"),
        ("GRAND TOTAL RAB", "='RAB Rekapitulasi'!G24"),
    ]

    for idx, (lbl, val) in enumerate(infos):
        r = 6 + idx
        ws_cover.row_dimensions[r].height = 20
        ws_cover.cell(row=r, column=1, value=lbl).font = font_card_lbl
        ws_cover.cell(row=r, column=1).alignment = align_left
        ws_cover.cell(row=r, column=1).fill = fill_card_bg
        ws_cover.cell(row=r, column=1).border = border_card
        
        ws_cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        c_val = ws_cover.cell(row=r, column=2, value=val)
        c_val.font = font_card_val
        c_val.alignment = align_left
        style_range(ws_cover, f"B{r}:C{r}", font=font_card_val, border=border_data)

    font_white_gt = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")

    for idx, (lbl, formula) in enumerate(summary_items):
        r = 6 + idx
        ws_cover.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
        c_lbl = ws_cover.cell(row=r, column=4, value=lbl)
        
        use_font = font_card_lbl if idx < 6 else font_white_gt
        use_fill = fill_card_bg if idx < 6 else fill_dark_emerald
        use_border = border_data if idx < 6 else border_grand
        
        c_lbl.font = use_font
        c_lbl.alignment = align_left
        c_lbl.fill = use_fill
        style_range(ws_cover, f"D{r}:E{r}", font=use_font, fill=use_fill, border=use_border)

        c_val = ws_cover.cell(row=r, column=6)
        c_val.value = formula
        c_val.font = use_font
        c_val.alignment = align_right
        c_val.fill = use_fill
        c_val.number_format = FMT_RP
        c_val.border = use_border

    # Signature Block
    sig_start = 18
    ws_cover.row_dimensions[sig_start].height = 18
    ws_cover.merge_cells(f"A{sig_start}:C{sig_start}")
    ws_cover[f"A{sig_start}"] = "Disetujui Oleh:"
    ws_cover[f"A{sig_start}"].font = font_sub_grp

    ws_cover.merge_cells(f"D{sig_start}:F{sig_start}")
    ws_cover[f"D{sig_start}"] = "Dibuat Oleh:"
    ws_cover[f"D{sig_start}"].font = font_sub_grp

    ws_cover.merge_cells(f"A{sig_start+1}:C{sig_start+1}")
    ws_cover[f"A{sig_start+1}"] = "Pejabat Pembuat Komitmen (PPK)"
    ws_cover[f"A{sig_start+1}"].font = font_reg_italic

    ws_cover.merge_cells(f"D{sig_start+1}:F{sig_start+1}")
    ws_cover[f"D{sig_start+1}"] = "Konsultan Perencana / Tim Teknik"
    ws_cover[f"D{sig_start+1}"].font = font_reg_italic

    for empty_r in range(sig_start+2, sig_start+5):
        ws_cover.row_dimensions[empty_r].height = 18

    ws_cover.merge_cells(f"A{sig_start+5}:C{sig_start+5}")
    ws_cover[f"A{sig_start+5}"] = "( _______________________________ )"
    ws_cover[f"A{sig_start+5}"].font = font_data_bold

    ws_cover.merge_cells(f"D{sig_start+5}:F{sig_start+5}")
    ws_cover[f"D{sig_start+5}"] = "( _______________________________ )"
    ws_cover[f"D{sig_start+5}"].font = font_data_bold

    ws_cover.column_dimensions['A'].width = 24
    ws_cover.column_dimensions['B'].width = 20
    ws_cover.column_dimensions['C'].width = 28
    ws_cover.column_dimensions['D'].width = 26
    ws_cover.column_dimensions['E'].width = 16
    ws_cover.column_dimensions['F'].width = 24


    # =========================================================================
    # SHEET 2: AHSP DETAIL
    # =========================================================================
    ws_ahsp = wb.create_sheet(title="AHSP Detail")
    ws_ahsp.views.sheetView[0].showGridLines = True

    ws_ahsp.merge_cells("A1:F1")
    ws_ahsp["A1"] = "ANALISIS HARGA SATUAN PEKERJAAN (AHSP) — PEMOTONGAN / PERAMPINGAN POHON"
    style_range(ws_ahsp, "A1:F1", font=font_title_main, fill=fill_dark_emerald, alignment=align_center_wrap, border=border_hdr)
    ws_ahsp.row_dimensions[1].height = 34

    ws_ahsp.merge_cells("A2:F2")
    ws_ahsp["A2"] = "Acuan: SE Ditjen BK No. 47/SE/Dk/2026  |  Overhead & Profit 15%  |  Satuan per Batang Pohon"
    style_range(ws_ahsp, "A2:F2", font=font_subtitle, fill=fill_medium_forest, alignment=align_center_wrap)
    ws_ahsp.row_dimensions[2].height = 22

    headers_ahsp = ["No.", "Uraian Komponen / Pekerjaan", "Sat.", "Koefisien", "H. Satuan (Rp)", "Jumlah (Rp)"]
    for c_idx, h_text in enumerate(headers_ahsp, 1):
        cell = ws_ahsp.cell(row=3, column=c_idx, value=h_text)
        cell.font = font_col_hdr
        cell.fill = fill_dark_emerald
        cell.alignment = align_center_wrap
        cell.border = border_hdr
    ws_ahsp.row_dimensions[3].height = 28

    # Data AHSP Definitions
    ahsp_groups = [
        {
            "group_label": "KELOMPOK A — PERAMPINGAN / PEMANGKASAN POHON (Pohon Tetap Berdiri / Hidup)",
            "items": [
                {
                    "label": "A.1 · Perampingan Ø 5–15 cm (Pohon Kecil / Perdu)",
                    "labor": [("Pekerja Biasa", "OH", 0.05, 120000), ("Tukang", "OH", 0.08, 160000), ("Mandor", "OH", 0.008, 185000)],
                    "equip": [("Chainsaw Kecil", "Jam", 0.3, 45000), ("Alat Bantu", "LS", 1.0, 750)],
                    "mat": [("Cat Luka Pohon", "Kaleng", 0.02, 95000)]
                },
                {
                    "label": "A.2 · Perampingan Ø 15–30 cm (Pohon Sedang Kecil)",
                    "labor": [("Pekerja Biasa", "OH", 0.08, 120000), ("Tukang", "OH", 0.12, 160000), ("Mandor", "OH", 0.01, 185000)],
                    "equip": [("Chainsaw Kecil", "Jam", 0.5, 45000), ("Alat Bantu", "LS", 1.0, 1125)],
                    "mat": [("Cat Luka Pohon", "Kaleng", 0.05, 95000)]
                },
                {
                    "label": "A.3 · Perampingan Ø 30–50 cm (Pohon Sedang)",
                    "labor": [("Pekerja Biasa", "OH", 0.15, 120000), ("Tukang", "OH", 0.2, 160000), ("Mandor", "OH", 0.015, 185000)],
                    "equip": [("Chainsaw Sedang", "Jam", 0.8, 65000), ("Alat Bantu", "LS", 1.0, 1500)],
                    "mat": [("Cat Luka Pohon", "Kaleng", 0.1, 95000)]
                },
                {
                    "label": "A.4 · Perampingan Ø 50–75 cm (Pohon Besar)",
                    "labor": [("Pekerja Biasa", "OH", 0.25, 120000), ("Tukang", "OH", 0.35, 160000), ("Mandor", "OH", 0.02, 185000)],
                    "equip": [("Chainsaw Sedang", "Jam", 1.2, 65000), ("Dump Truck - Angkut Limbah", "Jam", 0.5, 350000), ("Alat Bantu", "LS", 1.0, 1875)],
                    "mat": [("Cat Luka Pohon", "Kaleng", 0.15, 95000)]
                },
                {
                    "label": "A.5 · Perampingan Ø > 75 cm (Pohon Sangat Besar)",
                    "labor": [("Pekerja Biasa", "OH", 0.35, 120000), ("Tukang", "OH", 0.5, 160000), ("Mandor", "OH", 0.03, 185000)],
                    "equip": [("Chainsaw Besar", "Jam", 1.8, 90000), ("Dump Truck - Angkut Limbah", "Jam", 1.2, 350000), ("Alat Bantu", "LS", 1.0, 2625)],
                    "mat": [("Cat Luka Pohon", "Kaleng", 0.2, 95000)]
                },
                {
                    "label": "A.6 · Perampingan Ø > 100 cm (Pohon Raksasa)",
                    "labor": [("Pekerja Biasa", "OH", 0.50, 120000), ("Tukang", "OH", 0.70, 160000), ("Mandor", "OH", 0.04, 185000)],
                    "equip": [("Chainsaw Besar", "Jam", 2.5, 90000), ("Dump Truck - Angkut Limbah", "Jam", 1.8, 350000), ("Alat Bantu", "LS", 1.0, 3500)],
                    "mat": [("Cat Luka Pohon", "Kaleng", 0.25, 95000)]
                }
            ]
        },
        {
            "group_label": "KELOMPOK B — PEMOTONGAN / PENEBANGAN POHON (Ditebang Habis + Angkut Limbah)",
            "items": [
                {
                    "label": "B.1 · Pemotongan Ø < 15 cm (Pohon Kecil / Perdu)",
                    "labor": [("Pekerja Biasa", "OH", 0.06, 120000), ("Tukang", "OH", 0.08, 160000), ("Mandor", "OH", 0.008, 185000)],
                    "equip": [("Chainsaw Kecil", "Jam", 0.2, 45000), ("Alat Bantu", "LS", 1.0, 750)],
                    "mat": []
                },
                {
                    "label": "B.2 · Pemotongan Ø 15–30 cm (Pohon Sedang Kecil)",
                    "labor": [("Pekerja Biasa", "OH", 0.09, 120000), ("Tukang", "OH", 0.12, 160000), ("Mandor", "OH", 0.01, 185000)],
                    "equip": [("Chainsaw Kecil", "Jam", 0.5, 45000), ("Alat Bantu", "LS", 1.0, 1125)],
                    "mat": []
                },
                {
                    "label": "B.3 · Pemotongan Ø 30–50 cm (Pohon Sedang)",
                    "labor": [("Pekerja Biasa", "OH", 0.15, 120000), ("Tukang", "OH", 0.2, 160000), ("Mandor", "OH", 0.015, 185000)],
                    "equip": [("Chainsaw Sedang", "Jam", 0.8, 65000), ("Dump Truck - Angkut Limbah", "Jam", 0.3, 350000), ("Alat Bantu", "LS", 1.0, 1500)],
                    "mat": []
                },
                {
                    "label": "B.4 · Pemotongan Ø 50–75 cm (Pohon Besar)",
                    "labor": [("Pekerja Biasa", "OH", 0.25, 120000), ("Tukang", "OH", 0.35, 160000), ("Mandor", "OH", 0.02, 185000)],
                    "equip": [("Chainsaw Sedang", "Jam", 1.5, 65000), ("Dump Truck - Angkut Limbah", "Jam", 0.75, 350000), ("Alat Bantu", "LS", 1.0, 1875)],
                    "mat": []
                },
                {
                    "label": "B.5 · Pemotongan Ø 75–100 cm (Pohon Sangat Besar)",
                    "labor": [("Pekerja Biasa", "OH", 0.4, 120000), ("Tukang", "OH", 0.5, 160000), ("Mandor", "OH", 0.03, 185000)],
                    "equip": [("Chainsaw Besar", "Jam", 2.0, 90000), ("Dump Truck - Angkut Limbah", "Jam", 1.4, 350000), ("Alat Bantu", "LS", 1.0, 2250)],
                    "mat": []
                },
                {
                    "label": "B.6 · Pemotongan Ø > 100 cm (Pohon Raksasa)",
                    "labor": [("Pekerja Biasa", "OH", 0.6, 120000), ("Tukang", "OH", 0.75, 160000), ("Mandor", "OH", 0.05, 185000)],
                    "equip": [("Chainsaw Besar", "Jam", 3.0, 90000), ("Dump Truck - Angkut Limbah", "Jam", 2.35, 350000), ("Alat Bantu", "LS", 1.0, 3750)],
                    "mat": []
                }
            ]
        }
    ]

    curr_row = 4
    for group in ahsp_groups:
        ws_ahsp.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
        ws_ahsp.cell(row=curr_row, column=1, value=group["group_label"])
        style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", font=font_sec_hdr, fill=fill_light_sage, alignment=align_left, border=border_data)
        ws_ahsp.row_dimensions[curr_row].height = 24
        curr_row += 1

        for item in group["items"]:
            # Item Header
            ws_ahsp.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
            ws_ahsp.cell(row=curr_row, column=1, value=item["label"])
            style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", font=font_item_hdr, fill=fill_medium_forest, alignment=align_left, border=border_data)
            ws_ahsp.row_dimensions[curr_row].height = 22
            curr_row += 1

            sub_rows = []
            
            # Helper to write sub-category
            def write_sub_cat(sub_name, sub_data):
                nonlocal curr_row
                ws_ahsp.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
                ws_ahsp.cell(row=curr_row, column=1, value=sub_name)
                style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", font=font_sub_grp, fill=fill_mint_tint, alignment=align_left, border=border_data)
                ws_ahsp.row_dimensions[curr_row].height = 20
                curr_row += 1

                start_r = curr_row
                for name, sat, coef, price in sub_data:
                    ws_ahsp.cell(row=curr_row, column=1, value="")
                    ws_ahsp.cell(row=curr_row, column=2, value=name).font = font_data
                    ws_ahsp.cell(row=curr_row, column=3, value=sat).font = font_data
                    ws_ahsp.cell(row=curr_row, column=3).alignment = align_center

                    c_coef = ws_ahsp.cell(row=curr_row, column=4, value=coef)
                    c_coef.font = font_data
                    c_coef.alignment = align_center
                    c_coef.number_format = FMT_COEF

                    c_prc = ws_ahsp.cell(row=curr_row, column=5, value=price)
                    c_prc.font = font_data
                    c_prc.alignment = align_right
                    c_prc.number_format = FMT_RP

                    c_tot = ws_ahsp.cell(row=curr_row, column=6, value=f"=D{curr_row}*E{curr_row}")
                    c_tot.font = font_data
                    c_tot.alignment = align_right
                    c_tot.number_format = FMT_RP

                    style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", border=border_data)
                    ws_ahsp.row_dimensions[curr_row].height = 20
                    curr_row += 1

                end_r = curr_row - 1
                # Subtotal row
                ws_ahsp.cell(row=curr_row, column=1, value="")
                ws_ahsp.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=5)
                ws_ahsp.cell(row=curr_row, column=2, value=f"JUMLAH {sub_name}").font = font_subtotal
                ws_ahsp.cell(row=curr_row, column=2).alignment = align_left

                c_sub = ws_ahsp.cell(row=curr_row, column=6)
                if start_r <= end_r:
                    c_sub.value = f"=SUM(F{start_r}:F{end_r})"
                else:
                    c_sub.value = 0
                c_sub.font = font_subtotal
                c_sub.alignment = align_right
                c_sub.number_format = FMT_RP

                style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", fill=fill_zebra_light, border=border_subtotal)
                ws_ahsp.row_dimensions[curr_row].height = 20
                sub_r = curr_row
                curr_row += 1
                return sub_r

            r_labor = write_sub_cat("A. TENAGA KERJA", item["labor"])
            r_equip = write_sub_cat("B. PERALATAN", item["equip"])
            r_mat   = write_sub_cat("C. BAHAN / MATERIAL", item["mat"])

            # D. Jumlah Biaya Langsung
            ws_ahsp.cell(row=curr_row, column=1, value="D").font = font_direct_cost
            ws_ahsp.cell(row=curr_row, column=1).alignment = align_center
            ws_ahsp.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=5)
            ws_ahsp.cell(row=curr_row, column=2, value="JUMLAH BIAYA LANGSUNG  (A + B + C)").font = font_direct_cost
            
            c_dir = ws_ahsp.cell(row=curr_row, column=6, value=f"=F{r_labor}+F{r_equip}+F{r_mat}")
            c_dir.font = font_direct_cost
            c_dir.alignment = align_right
            c_dir.number_format = FMT_RP

            style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", fill=fill_light_sage, border=border_subtotal)
            ws_ahsp.row_dimensions[curr_row].height = 22
            r_dir = curr_row
            curr_row += 1

            # E. Overhead & Profit 15%
            ws_ahsp.cell(row=curr_row, column=1, value="E").font = font_overhead
            ws_ahsp.cell(row=curr_row, column=1).alignment = align_center
            ws_ahsp.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=5)
            ws_ahsp.cell(row=curr_row, column=2, value="OVERHEAD & PROFIT  (15% x D)").font = font_overhead
            
            c_ovh = ws_ahsp.cell(row=curr_row, column=6, value=f"=F{r_dir}*0.15")
            c_ovh.font = font_overhead
            c_ovh.alignment = align_right
            c_ovh.number_format = FMT_RP

            style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", border=border_data)
            ws_ahsp.row_dimensions[curr_row].height = 20
            r_ovh = curr_row
            curr_row += 1

            # F. Harga Satuan Pekerjaan (per Batang)
            ws_ahsp.cell(row=curr_row, column=1, value="F").font = font_grand_unit
            ws_ahsp.cell(row=curr_row, column=1).alignment = align_center
            ws_ahsp.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=5)
            ws_ahsp.cell(row=curr_row, column=2, value="HARGA SATUAN PEKERJAAN  (D + E)  — per Batang").font = font_grand_unit
            
            c_tot_unit = ws_ahsp.cell(row=curr_row, column=6, value=f"=F{r_dir}+F{r_ovh}")
            c_tot_unit.font = font_grand_unit
            c_tot_unit.alignment = align_right
            c_tot_unit.number_format = FMT_RP

            style_range(ws_ahsp, f"A{curr_row}:F{curr_row}", fill=fill_dark_emerald, border=border_grand)
            ws_ahsp.row_dimensions[curr_row].height = 24
            curr_row += 2  # spacing

    ws_ahsp.column_dimensions['A'].width = 6
    ws_ahsp.column_dimensions['B'].width = 46
    ws_ahsp.column_dimensions['C'].width = 10
    ws_ahsp.column_dimensions['D'].width = 14
    ws_ahsp.column_dimensions['E'].width = 18
    ws_ahsp.column_dimensions['F'].width = 20


    # =========================================================================
    # SHEET 3: RAB REKAPITULASI
    # =========================================================================
    ws_rab = wb.create_sheet(title="RAB Rekapitulasi")
    ws_rab.views.sheetView[0].showGridLines = True

    ws_rab.merge_cells("A1:G1")
    ws_rab["A1"] = "REKAPITULASI RENCANA ANGGARAN BIAYA (RAB)"
    style_range(ws_rab, "A1:G1", font=font_title_main, fill=fill_dark_emerald, alignment=align_center_wrap, border=border_hdr)
    ws_rab.row_dimensions[1].height = 34

    ws_rab.merge_cells("A2:G2")
    ws_rab["A2"] = "Pekerjaan Pemotongan / Perampingan Pohon  |  AHSP 2026 PUPR  |  SE Ditjen BK No. 47/SE/Dk/2026"
    style_range(ws_rab, "A2:G2", font=font_subtitle, fill=fill_medium_forest, alignment=align_center_wrap)
    ws_rab.row_dimensions[2].height = 22

    headers_rab = ["No.", "Kode", "Uraian Pekerjaan", "Satuan", "Volume\n(Batang)", "Harga Satuan\n(Rp)", "Jumlah Harga\n(Rp)"]
    for c_idx, h_text in enumerate(headers_rab, 1):
        cell = ws_rab.cell(row=3, column=c_idx, value=h_text)
        cell.font = font_col_hdr
        cell.fill = fill_dark_emerald
        cell.alignment = align_center_wrap
        cell.border = border_hdr
    ws_rab.row_dimensions[3].height = 30

    # Kelompok A Rows
    ws_rab.merge_cells("A4:G4")
    ws_rab["A4"] = "A.  PERAMPINGAN / PEMANGKASAN POHON  (Pohon Tetap Berdiri / Hidup)"
    style_range(ws_rab, "A4:G4", font=font_sec_hdr, fill=fill_light_sage, alignment=align_left, border=border_data)
    ws_rab.row_dimensions[4].height = 24

    items_a = [
        (1, "A.1", "Perampingan Ø 5–15 cm — Pohon kecil / perdu", "Batang", 3, 42000),
        (2, "A.2", "Perampingan Ø 15–30 cm — Pohon sedang kecil", "Batang", 5, 68000),
        (3, "A.3", "Perampingan Ø 30–50 cm — Pohon sedang", "Batang", 0, 133500),
        (4, "A.4", "Perampingan Ø 50–75 cm — Pohon besar", "Batang", 0, 413000),
        (5, "A.5", "Perampingan Ø > 75 cm — Pohon sangat besar", "Batang", 0, 841000),
        (6, "A.6", "Perampingan Ø > 100 cm — Pohon raksasa", "Batang", 0, 1221000),
    ]

    for idx, (no, kode, uraian, sat, vol, price) in enumerate(items_a):
        r = 5 + idx
        ws_rab.cell(row=r, column=1, value=no).alignment = align_center
        ws_rab.cell(row=r, column=2, value=kode).alignment = align_center
        ws_rab.cell(row=r, column=3, value=uraian).alignment = align_left
        ws_rab.cell(row=r, column=4, value=sat).alignment = align_center

        c_vol = ws_rab.cell(row=r, column=5, value=vol)
        c_vol.alignment = align_center
        c_vol.number_format = FMT_INT

        c_prc = ws_rab.cell(row=r, column=6, value=price)
        c_prc.alignment = align_right
        c_prc.number_format = FMT_RP

        c_tot = ws_rab.cell(row=r, column=7, value=f"=E{r}*F{r}")
        c_tot.alignment = align_right
        c_tot.number_format = FMT_RP

        fill_r = fill_zebra_light if idx % 2 == 1 else None
        style_range(ws_rab, f"A{r}:G{r}", font=font_data, fill=fill_r, border=border_data)
        ws_rab.row_dimensions[r].height = 20

    # Subtotal A
    ws_rab.merge_cells("A11:F11")
    ws_rab["A11"] = "Sub-Total Kelompok A (Perampingan / Pemangkasan)"
    ws_rab["A11"].alignment = align_left
    c_suba = ws_rab.cell(row=11, column=7, value="=SUM(G5:G10)")
    c_suba.alignment = align_right
    c_suba.number_format = FMT_RP
    style_range(ws_rab, "A11:G11", font=font_subtotal, fill=fill_mint_tint, border=border_subtotal)
    ws_rab.row_dimensions[11].height = 22

    # Kelompok B Rows
    ws_rab.merge_cells("A12:G12")
    ws_rab["A12"] = "B.  PEMOTONGAN / PENEBANGAN POHON  (Ditebang Habis + Angkut Limbah)"
    style_range(ws_rab, "A12:G12", font=font_sec_hdr, fill=fill_light_sage, alignment=align_left, border=border_data)
    ws_rab.row_dimensions[12].height = 24

    items_b = [
        (1, "B.1", "Pemotongan Ø < 15 cm — Pohon kecil / perdu", "Batang", 8, 36000),
        (2, "B.2", "Pemotongan Ø 15–30 cm — Pohon sedang kecil", "Batang", 0, 64000),
        (3, "B.3", "Pemotongan Ø 30–50 cm — Pohon sedang", "Batang", 3, 243000),
        (4, "B.4", "Pemotongan Ø 50–75 cm — Pohon besar", "Batang", 2, 519500),
        (5, "B.5", "Pemotongan Ø 75–100 cm — Pohon sangat besar", "Batang", 2, 927000),
        (6, "B.6", "Pemotongan Ø > 100 cm — Pohon raksasa", "Batang", 1, 1492500),
    ]

    for idx, (no, kode, uraian, sat, vol, price) in enumerate(items_b):
        r = 13 + idx
        ws_rab.cell(row=r, column=1, value=no).alignment = align_center
        ws_rab.cell(row=r, column=2, value=kode).alignment = align_center
        ws_rab.cell(row=r, column=3, value=uraian).alignment = align_left
        ws_rab.cell(row=r, column=4, value=sat).alignment = align_center

        c_vol = ws_rab.cell(row=r, column=5, value=vol)
        c_vol.alignment = align_center
        c_vol.number_format = FMT_INT

        c_prc = ws_rab.cell(row=r, column=6, value=price)
        c_prc.alignment = align_right
        c_prc.number_format = FMT_RP

        c_tot = ws_rab.cell(row=r, column=7, value=f"=E{r}*F{r}")
        c_tot.alignment = align_right
        c_tot.number_format = FMT_RP

        fill_r = fill_zebra_light if idx % 2 == 1 else None
        style_range(ws_rab, f"A{r}:G{r}", font=font_data, fill=fill_r, border=border_data)
        ws_rab.row_dimensions[r].height = 20

    # Subtotal B
    ws_rab.merge_cells("A19:F19")
    ws_rab["A19"] = "Sub-Total Kelompok B (Pemotongan / Penebangan)"
    ws_rab["A19"].alignment = align_left
    c_subb = ws_rab.cell(row=19, column=7, value="=SUM(G13:G18)")
    c_subb.alignment = align_right
    c_subb.number_format = FMT_RP
    style_range(ws_rab, "A19:G19", font=font_subtotal, fill=fill_mint_tint, border=border_subtotal)
    ws_rab.row_dimensions[19].height = 22

    # Recap Calculations
    recap_rows = [
        (20, "JUMLAH BIAYA LANGSUNG FISIK  (A + B)", "=G11+G19", font_direct_cost, fill_light_sage, border_subtotal, 22),
        (21, "BIAYA PENERAPAN SMKK / K3  (2.5% x Total Fisik)", "=G20*0.025", font_data_bold, fill_zebra_light, border_data, 20),
        (22, "JUMLAH BIAYA FISIK + SMKK", "=G20+G21", font_direct_cost, fill_light_sage, border_subtotal, 22),
        (23, "PAJAK PERTAMBAHAN NILAI  (PPN 11%)", "=G22*0.11", font_data_bold, fill_zebra_light, border_data, 20),
        (24, "GRAND TOTAL RENCANA ANGGARAN BIAYA (RAB)", "=G22+G23", font_grand_total, fill_dark_emerald, border_grand, 26),
    ]

    for r_num, label, formula, font_style, fill_style, border_style, height in recap_rows:
        ws_rab.merge_cells(start_row=r_num, start_column=1, end_row=r_num, end_column=6)
        ws_rab.cell(row=r_num, column=1, value=label).alignment = align_left
        
        c_val = ws_rab.cell(row=r_num, column=7, value=formula)
        c_val.alignment = align_right
        c_val.number_format = FMT_RP

        style_range(ws_rab, f"A{r_num}:G{r_num}", font=font_style, fill=fill_style, border=border_style)
        ws_rab.row_dimensions[r_num].height = height

    # Note Box
    ws_rab.row_dimensions[25].height = 14

    ws_rab.merge_cells("A26:G26")
    ws_rab["A26"] = "📌 CATATAN TEKNIS & SYARAT PELAKSANAAN"
    style_range(ws_rab, "A26:G26", font=font_sub_grp, fill=fill_mint_tint, border=border_data)
    ws_rab.row_dimensions[26].height = 20

    notes_list = [
        "1. Regulasi Acuan: SE Ditjen Bina Konstruksi No. 47/SE/Dk/2026 (AHSP PUPR 2026).",
        "2. Harga Satuan Dasar (HSD): Upah dan sewa alat mengacu pada acuan HSD standar. Wajib disesuaikan HSD Pemda setempat untuk tender resmi.",
        "3. Penerapan SMKK (K3): Biaya K3 Konstruksi sebesar 2,5% diperhitungkan sesuai Permen PUPR No. 10/2021.",
        "4. Pajak Pertambahan Nilai: PPN 11% dihitung dari Total Fisik + SMKK sesuai UU HPP No. 7/2021.",
        "5. Komponen Angkut Limbah: Limbah kayu/ranting diangkut menggunakan armada Dump Truck (Angkut Limbah) untuk menjaga kebersihan lokasi kerja.",
    ]

    for idx, note in enumerate(notes_list):
        r = 27 + idx
        ws_rab.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws_rab.cell(row=r, column=1, value=note).font = font_card_val
        ws_rab.cell(row=r, column=1).alignment = align_left
        style_range(ws_rab, f"A{r}:G{r}", fill=fill_card_bg, border=border_data)
        ws_rab.row_dimensions[r].height = 18

    ws_rab.column_dimensions['A'].width = 6
    ws_rab.column_dimensions['B'].width = 10
    ws_rab.column_dimensions['C'].width = 52
    ws_rab.column_dimensions['D'].width = 10
    ws_rab.column_dimensions['E'].width = 16
    ws_rab.column_dimensions['F'].width = 22
    ws_rab.column_dimensions['G'].width = 24

    # Save to disk with fallback if open in Excel
    try:
        wb.save(output_path)
        print(f"File successfully written to: {output_path}")
    except PermissionError:
        alt_path = output_path.replace(".xlsx", "_Terbaru.xlsx")
        wb.save(alt_path)
        print(f"Permission denied for {output_path} (file open). Saved to fallback: {alt_path}")

if __name__ == "__main__":
    create_styled_rab_excel(r"c:\Users\Windows 11\Documents\RAB DAN ANALISA PEMOTONGAN POHON\RAB_Pemotongan_Pohon_AHSP2026(tanpa wood chipper).xlsx")
    create_styled_rab_excel(r"c:\Users\Windows 11\Documents\RAB DAN ANALISA PEMOTONGAN POHON\RAB_Pemotongan_Pohon_AHSP2026.xlsx")
